# ==============================
# FINAL PERSONAL FINANCE MANAGER
# ==============================

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv, os
from datetime import datetime
from collections import defaultdict

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, Reference

# ---------- CONFIG ----------
DATA_FILE = "finance_data.csv"

BUDGETS = {
    "Food": 5000,
    "Shopping": 3000,
    "Travel": 4000,
    "Bills": 7000
}

THEMES = {
    "light": {"bg": "#f5f5f5", "fg": "#000000"},
    "dark": {"bg": "#1e1e1e", "fg": "#ffffff"}
}

current_theme = "light"

# ---------- INIT CSV ----------
if not os.path.exists(DATA_FILE):
    with open(DATA_FILE, "w", newline="") as f:
        csv.writer(f).writerow(["Date", "Type", "Category", "Amount", "Note"])

# ---------- FUNCTIONS ----------
def read_data():
    data = []
    with open(DATA_FILE) as f:
        for r in csv.DictReader(f):
            r["Amount"] = float(r["Amount"])
            data.append(r)
    return data

def add_entry():
    try:
        amt = float(amount_entry.get())
        date = date_entry.get()
        datetime.strptime(date, "%Y-%m-%d")
    except:
        messagebox.showerror("Error", "Invalid input")
        return

    with open(DATA_FILE, "a", newline="") as f:
        csv.writer(f).writerow([
            date,
            type_var.get(),
            category_entry.get(),
            amt,
            note_entry.get()
        ])

    clear_inputs()
    refresh()

def clear_inputs():
    for e in (date_entry, category_entry, amount_entry, note_entry):
        e.delete(0, tk.END)

def refresh():
    data = read_data()
    month = month_var.get()

    if month != "All":
        data = [d for d in data if d["Date"].startswith(month)]

    total_income = sum(d["Amount"] for d in data if d["Type"] == "Income")
    total_expense = sum(d["Amount"] for d in data if d["Type"] == "Expense")

    income_lbl.config(text=f"Total Income: ₹{total_income}")
    expense_lbl.config(text=f"Total Expense: ₹{total_expense}")
    balance_lbl.config(text=f"Balance: ₹{total_income-total_expense}")

    show_alerts(data)
    plot_pie(data)
    plot_trend(data)

def show_alerts(data):
    alerts = ""
    spent = defaultdict(float)

    for d in data:
        if d["Type"] == "Expense":
            spent[d["Category"]] += d["Amount"]

    for c,v in spent.items():
        if c in BUDGETS and v > BUDGETS[c]:
            alerts += f"⚠ {c} exceeded budget\n"

    alert_lbl.config(text=alerts)

def plot_pie(data):
    pie_fig.clear()
    ax = pie_fig.add_subplot(111)

    cat = defaultdict(float)
    for d in data:
        if d["Type"] == "Expense":
            cat[d["Category"]] += d["Amount"]

    if cat:
        ax.pie(cat.values(), labels=cat.keys(), autopct="%1.1f%%")
        ax.set_title("Expenses by Category")
    else:
        ax.text(0.5,0.5,"No data",ha="center",va="center")
        ax.axis("off")

    pie_canvas.draw()

def plot_trend(data):
    trend_fig.clear()
    ax = trend_fig.add_subplot(111)

    inc, exp = defaultdict(float), defaultdict(float)

    for d in data:
        m = d["Date"][:7]
        if d["Type"] == "Income": inc[m]+=d["Amount"]
        else: exp[m]+=d["Amount"]

    months = sorted(set(inc)|set(exp))
    if months:
        ax.plot(months,[inc[m] for m in months],label="Income",marker="o")
        ax.plot(months,[exp[m] for m in months],label="Expense",marker="o")
        ax.legend()
        ax.grid(True)
    else:
        ax.text(0.5,0.5,"No data",ha="center",va="center")
        ax.axis("off")

    ax.set_title("Monthly Trend")
    trend_canvas.draw()

def export_excel():
    data = read_data()
    if not data:
        return

    path = filedialog.asksaveasfilename(defaultextension=".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Finance"

    headers = ["Date","Type","Category","Amount","Note"]
    for i,h in enumerate(headers,1):
        c = ws.cell(row=1,column=i,value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFD966")

    for r,d in enumerate(data,2):
        ws.append([d[h] for h in headers])

    # Excel Pie Chart
    pie = PieChart()
    pie.title = "Expenses"
    cats = defaultdict(float)
    for d in data:
        if d["Type"]=="Expense":
            cats[d["Category"]]+=d["Amount"]

    row = len(data)+3
    ws.cell(row=row,column=1,value="Category")
    ws.cell(row=row,column=2,value="Amount")

    for i,(c,v) in enumerate(cats.items(),row+1):
        ws.cell(row=i,column=1,value=c)
        ws.cell(row=i,column=2,value=v)

    pie.add_data(Reference(ws,row+1,row+len(cats),2,2))
    pie.set_categories(Reference(ws,row+1,row+len(cats),1,1))
    ws.add_chart(pie,"G2")

    wb.save(path)
    messagebox.showinfo("Done","Excel exported successfully")

def toggle_theme():
    global current_theme
    current_theme = "dark" if current_theme=="light" else "light"
    apply_theme()

def apply_theme():
    t = THEMES[current_theme]
    root.configure(bg=t["bg"])
    for w in root.winfo_children():
        try:
            w.configure(bg=t["bg"], fg=t["fg"])
        except:
            pass

# ---------- UI ----------
root = tk.Tk()
root.title("Personal Finance Manager – FINAL")
root.geometry("1100x720")

# Inputs
top = tk.Frame(root)
top.pack(fill="x",padx=10,pady=5)

tk.Label(top,text="Date").grid(row=0,column=0)
date_entry = tk.Entry(top); date_entry.grid(row=0,column=1)

type_var = tk.StringVar(value="Expense")
tk.Radiobutton(top,text="Expense",variable=type_var,value="Expense").grid(row=0,column=2)
tk.Radiobutton(top,text="Income",variable=type_var,value="Income").grid(row=0,column=3)

tk.Label(top,text="Category").grid(row=1,column=0)
category_entry = tk.Entry(top); category_entry.grid(row=1,column=1)

tk.Label(top,text="Amount").grid(row=1,column=2)
amount_entry = tk.Entry(top); amount_entry.grid(row=1,column=3)

tk.Label(top,text="Note").grid(row=2,column=0)
note_entry = tk.Entry(top,width=40); note_entry.grid(row=2,column=1,columnspan=3)

tk.Button(top,text="Add Entry",command=add_entry,bg="green",fg="white").grid(row=3,column=0,columnspan=4,pady=5)

# Controls
ctrl = tk.Frame(root)
ctrl.pack(fill="x")

month_var = tk.StringVar(value="All")
months = sorted({d["Date"][:7] for d in read_data()})
ttk.Combobox(ctrl,textvariable=month_var,values=["All"]+months,width=10).pack(side="left",padx=5)
tk.Button(ctrl,text="Refresh",command=refresh).pack(side="left")
tk.Button(ctrl,text="Export Excel",command=export_excel).pack(side="left",padx=5)
tk.Button(ctrl,text="Toggle Theme",command=toggle_theme).pack(side="left")

# Summary
summary = tk.Frame(root)
summary.pack(fill="x",padx=10)

income_lbl = tk.Label(summary)
expense_lbl = tk.Label(summary)
balance_lbl = tk.Label(summary,font=("Arial",11,"bold"))
alert_lbl = tk.Label(summary,fg="red")

income_lbl.pack(anchor="w")
expense_lbl.pack(anchor="w")
balance_lbl.pack(anchor="w")
alert_lbl.pack(anchor="w")

# Charts
charts = tk.Frame(root)
charts.pack(fill="both",expand=True)

pie_fig = plt.Figure(figsize=(5,4))
pie_canvas = FigureCanvasTkAgg(pie_fig,charts)
pie_canvas.get_tk_widget().pack(side="left",fill="both",expand=True)

trend_fig = plt.Figure(figsize=(5,4))
trend_canvas = FigureCanvasTkAgg(trend_fig,charts)
trend_canvas.get_tk_widget().pack(side="left",fill="both",expand=True)

apply_theme()
refresh()
root.mainloop()
